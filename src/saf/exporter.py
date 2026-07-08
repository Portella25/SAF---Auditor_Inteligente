from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CORES_STATUS = {
    "Aprovado": "C6EFCE",
    "Atenção": "FFEB9C",
    "Atencao": "FFEB9C",
    "Rejeitado": "FFC7CE",
}


def exportar_relatorio_contestacao(auditoria_df: pd.DataFrame, resumo_df: pd.DataFrame, caminho_saida: str | Path) -> Path:
    caminho = Path(caminho_saida)
    caminho.parent.mkdir(parents=True, exist_ok=True)

    contestacao = auditoria_df[auditoria_df["status"] != "Aprovado"].copy()
    if contestacao.empty:
        contestacao = auditoria_df.head(0).copy()

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        resumo_df.to_excel(writer, index=False, sheet_name="Resumo Executivo")
        auditoria_df.to_excel(writer, index=False, sheet_name="Auditoria Completa")
        contestacao.to_excel(writer, index=False, sheet_name="Contestacao")

        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            for celula in sheet[1]:
                celula.font = Font(bold=True, color="FFFFFF")
                celula.fill = PatternFill("solid", fgColor="1F4E78")
                celula.alignment = Alignment(horizontal="center")
            for coluna in sheet.columns:
                largura = max(len(str(celula.value or "")) for celula in coluna)
                letra = get_column_letter(coluna[0].column)
                sheet.column_dimensions[letra].width = min(max(largura + 2, 12), 58)

        for nome_aba in ["Auditoria Completa", "Contestacao"]:
            sheet = writer.book[nome_aba]
            cabecalhos = [celula.value for celula in sheet[1]]
            if "status" not in cabecalhos:
                continue
            posicao_status = cabecalhos.index("status") + 1
            for linha in range(2, sheet.max_row + 1):
                status = sheet.cell(row=linha, column=posicao_status).value
                cor = CORES_STATUS.get(status)
                if cor:
                    for coluna in range(1, sheet.max_column + 1):
                        sheet.cell(row=linha, column=coluna).fill = PatternFill("solid", fgColor=cor)

    return caminho
